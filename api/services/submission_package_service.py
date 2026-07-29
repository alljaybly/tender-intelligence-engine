"""
Enterprise Submission Package Generator
========================================
Generates a single ZIP file containing all submission documents.

Primary button: "Generate Submission Package"

Package contents:
  01 Executive Summary.pdf
  02 Tender Completion Guide.pdf
  03 Tender Readiness Assessment.pdf
  04 Submission Letter.pdf
  05 Bid Response Roadmap.pdf
  06 Tender Integrity Audit.pdf
  07 Processing Audit.pdf
  08 Evidence Report.pdf
  09 BOQ.xlsx
  10 BOQ.csv
  11 Metadata.json
  12 Original Tender.pdf (if available)
  PACKAGE_MANIFEST.txt

All reports use common branding, headers, footers, page numbering, version numbering.
"""
import io
import json
import logging
import os
import zipfile
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable, KeepTogether, ListFlowable, ListItem,
)

from .confidence_service import compute_composite_confidence
from .summary_builder import build_clean_summary
from .submission_letter_service import generate_submission_letter
from .tender_completion_guide import generate_completion_guide
from .tender_readiness_service import generate_readiness_pdf
from .roadmap_audit_generator import generate_roadmap_pdf, generate_audit_pdf
from .evidence_service import build_evidence_system, build_evidence_pdf_sections
from .report_framework import (
    build_professional_report, get_pdf_styles, build_pdf_sections,
    VERIFICATION_NOTICE, VERIFICATION_NOTICE_FULL,
)
from .tender_readiness_service import build_readiness_report

logger = logging.getLogger(__name__)

PACKAGE_VERSION = "2.0.0"

# ── Color palette (enterprise, consistent across all reports) ─────────
PRIMARY_BLUE = colors.HexColor("#1F4E79")
ACCENT_BLUE = colors.HexColor("#2B78AE")
LIGHT_BLUE_BG = colors.HexColor("#E8F0FE")
VERY_LIGHT_BG = colors.HexColor("#F5F8FC")
WARNING_AMBER = colors.HexColor("#FFF3CD")
ERROR_RED = colors.HexColor("#F8D7DA")
SUCCESS_GREEN = colors.HexColor("#D4EDDA")
TEXT_DARK = colors.HexColor("#222222")
TEXT_MEDIUM = colors.HexColor("#555555")
TEXT_LIGHT = colors.HexColor("#888888")
WHITE = colors.white

# ── Page dimensions ───────────────────────────────────────────────────
PAGE_WIDTH, PAGE_HEIGHT = A4
MARGIN = 2 * cm
TOP_MARGIN = 2.5 * cm
BOTTOM_MARGIN = 2 * cm

# ── Styles ────────────────────────────────────────────────────────────
styles = getSampleStyleSheet()

COVER_TITLE_STYLE = ParagraphStyle(
    "CoverTitle", parent=styles["Title"],
    fontSize=28, leading=34, textColor=PRIMARY_BLUE,
    spaceAfter=6, alignment=TA_CENTER,
)
COVER_SUBTITLE_STYLE = ParagraphStyle(
    "CoverSubtitle", parent=styles["Normal"],
    fontSize=14, leading=18, textColor=TEXT_MEDIUM,
    spaceAfter=30, alignment=TA_CENTER,
)
SECTION_STYLE = ParagraphStyle(
    "SectionHeader", parent=styles["Heading1"],
    fontSize=18, leading=24, textColor=PRIMARY_BLUE,
    spaceBefore=20, spaceAfter=12,
)
SUBSECTION_STYLE = ParagraphStyle(
    "SubSectionHeader", parent=styles["Heading2"],
    fontSize=13, leading=17, textColor=PRIMARY_BLUE,
    spaceBefore=14, spaceAfter=8,
)
BODY_STYLE = ParagraphStyle(
    "PackageBody", parent=styles["Normal"],
    fontSize=9.5, leading=14, textColor=TEXT_DARK,
    spaceAfter=4, alignment=TA_JUSTIFY,
)
BODY_BOLD_STYLE = ParagraphStyle(
    "PackageBodyBold", parent=BODY_STYLE,
    fontName="Helvetica-Bold",
)
FOOTER_STYLE = ParagraphStyle(
    "Footer", parent=styles["Normal"],
    fontSize=7, leading=9, textColor=TEXT_LIGHT,
    alignment=TA_CENTER,
)
TOC_STYLE = ParagraphStyle(
    "TOCItem", parent=styles["Normal"],
    fontSize=11, leading=18, textColor=TEXT_DARK,
    leftIndent=20, spaceAfter=4,
)
TOC_SECTION_STYLE = ParagraphStyle(
    "TOCSection", parent=styles["Normal"],
    fontSize=12, leading=20, textColor=PRIMARY_BLUE,
    leftIndent=10, spaceAfter=6, fontName="Helvetica-Bold",
)


# ═══════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════

def _hr() -> HRFlowable:
    return HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC"),
                       spaceBefore=6, spaceAfter=6)


def _value_or_unavailable(value: Any, label: str = "") -> str:
    if value is None or value == "" or value == "N/A":
        if label:
            return f"[Unavailable — {label}]"
        return "[Unavailable]"
    return str(value)


def _format_currency(value: Any, currency_symbol: str = "R") -> str:
    if value is None:
        return "—"
    try:
        return f"{currency_symbol} {float(value):,.2f}"
    except (ValueError, TypeError):
        return str(value)


def _make_table(headers: List[str], rows: List[List[str]]) -> Table:
    data = [headers] + rows
    table = Table(data, colWidths=None, hAlign="LEFT")
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_commands.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8F9FA")))
    table.setStyle(TableStyle(style_commands))
    return table


def _build_pdf_document(title: str, story: List, job_id: str) -> BytesIO:
    """Build a PDF document with standard header/footer/page numbering."""
    buffer = BytesIO()

    def _header_footer(canvas, doc):
        canvas.saveState()
        # Header
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(TEXT_LIGHT)
        canvas.drawString(MARGIN, PAGE_HEIGHT - 1.5 * cm,
                          f"Tender Engine | {title}")
        canvas.drawRightString(PAGE_WIDTH - MARGIN, PAGE_HEIGHT - 1.5 * cm,
                               f"Job: {job_id[:12]}...")
        # Header line
        canvas.setStrokeColor(PRIMARY_BLUE)
        canvas.setLineWidth(0.5)
        canvas.line(MARGIN, PAGE_HEIGHT - 1.7 * cm,
                    PAGE_WIDTH - MARGIN, PAGE_HEIGHT - 1.7 * cm)
        # Footer
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(TEXT_LIGHT)
        canvas.drawCentredString(PAGE_WIDTH / 2, 1.2 * cm,
                                 f"Page {doc.page}")
        canvas.drawString(MARGIN, 1.2 * cm,
                          f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(PAGE_WIDTH - MARGIN, 1.2 * cm,
                               f"v{PACKAGE_VERSION}")
        # Footer line
        canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
        canvas.setLineWidth(0.3)
        canvas.line(MARGIN, 1.5 * cm, PAGE_WIDTH - MARGIN, 1.5 * cm)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=TOP_MARGIN, bottomMargin=BOTTOM_MARGIN,
        title=title, author="Tender Engine",
    )
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════
# Cover Page
# ═══════════════════════════════════════════════════════════════════════

def _build_cover_page(story: List, result_data: Dict[str, Any], job_id: str) -> None:
    """Build the cover page with branding."""
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph("SUBMISSION PACKAGE", COVER_TITLE_STYLE))
    story.append(Paragraph(
        "Enterprise Tender Response Document",
        COVER_SUBTITLE_STYLE,
    ))
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="60%", thickness=2, color=PRIMARY_BLUE, spaceAfter=8, spaceBefore=4))
    story.append(Spacer(1, 1 * cm))

    filename = result_data.get("filename", "Unknown Document")
    story.append(Paragraph(f"<b>{filename}</b>", ParagraphStyle(
        "CoverFilename", parent=BODY_STYLE, fontSize=12, leading=16, alignment=TA_CENTER,
    )))
    story.append(Spacer(1, 0.5 * cm))

    cover_data = [
        ["Job ID", job_id[:16] + "..." if len(job_id) > 16 else job_id],
        ["Sector", _value_or_unavailable(result_data.get("detected_sector"))],
        ["Location", ", ".join(result_data.get("detected_locations", []) or []) or "Not detected"],
        ["Date Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Package Version", PACKAGE_VERSION],
        ["Pipeline Version", _value_or_unavailable(result_data.get("pipeline_version"))],
        ["Status", result_data.get("status", "unknown").replace("_", " ").title()],
    ]
    cover_table = Table(cover_data, colWidths=[4.5 * cm, 10 * cm])
    cover_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), PRIMARY_BLUE),
        ("TEXTCOLOR", (1, 0), (1, -1), TEXT_DARK),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#EEEEEE")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
    ]))
    story.append(cover_table)

    story.append(Spacer(1, 1.5 * cm))
    story.append(Paragraph(
        "CONFIDENTIAL — This submission package is automatically generated "
        "from processed tender data. All confidence levels and warnings are "
        "preserved transparently.",
        ParagraphStyle("CoverDisclaimer", parent=BODY_STYLE,
                       fontSize=7, textColor=TEXT_LIGHT, fontName="Helvetica-Oblique",
                       alignment=TA_CENTER),
    ))
    story.append(PageBreak())


# ═══════════════════════════════════════════════════════════════════════
# Table of Contents
# ═══════════════════════════════════════════════════════════════════════

def _build_table_of_contents(story: List) -> None:
    """Build the table of contents page."""
    story.append(Spacer(1, 2 * cm))
    story.append(Paragraph("TABLE OF CONTENTS", COVER_TITLE_STYLE))
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="60%", thickness=1, color=PRIMARY_BLUE, spaceAfter=12, spaceBefore=4))
    story.append(Spacer(1, 0.5 * cm))

    toc_items = [
        ("01", "Executive Summary", "Summary of extracted data, completeness, and key findings"),
        ("02", "Tender Readiness Assessment", "Assessment of document completeness and readiness for submission"),
        ("03", "Submission Letter", "Formal submission letter for bid response"),
        ("04", "Bid Response Roadmap", "Structured guide with actionable steps for bid submission"),
        ("05", "Tender Integrity Audit", "Audit trail of all processing stages with evidence"),
        ("06", "Processing Audit", "Detailed processing log with OCR, currency, BOQ, and entity classification"),
        ("07", "BOQ (Excel)", "Bill of Quantities in Excel format"),
        ("08", "BOQ (CSV)", "Bill of Quantities in CSV format"),
        ("09", "Metadata", "Complete processing metadata in JSON format"),
        ("10", "Original Tender", "Original uploaded tender document (if available)"),
    ]

    for num, title, desc in toc_items:
        story.append(Paragraph(
            f"<b>{num}</b> &nbsp;&nbsp; {title}",
            TOC_SECTION_STYLE,
        ))
        story.append(Paragraph(
            f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; {desc}",
            TOC_STYLE,
        ))
        story.append(Spacer(1, 0.2 * cm))

    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="60%", thickness=0.5, color=colors.HexColor("#CCCCCC"),
                             spaceAfter=8, spaceBefore=4))
    story.append(Paragraph(
        f"Package Version: {PACKAGE_VERSION} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle("TOCFooter", parent=BODY_STYLE, fontSize=8, textColor=TEXT_MEDIUM, alignment=TA_CENTER),
    ))
    story.append(PageBreak())


# ═══════════════════════════════════════════════════════════════════════
# Build Metadata JSON
# ═══════════════════════════════════════════════════════════════════════

def _build_metadata(result_data: Dict[str, Any], job_id: str,
                    evidence_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Build comprehensive metadata JSON."""
    currency_data = result_data.get("detected_currency", {})
    if isinstance(currency_data, dict):
        currency_code = currency_data.get("currency_code")
        currency_conf = currency_data.get("confidence", 0)
        currency_method = currency_data.get("detection_method", "unknown")
    else:
        currency_code = None
        currency_conf = 0
        currency_method = "unknown"

    metadata = {
        "package_version": PACKAGE_VERSION,
        "generated_at": datetime.now().isoformat(),
        "job_id": job_id,
        "filename": result_data.get("filename", "Unknown"),
        "status": result_data.get("status", "unknown"),
        "pipeline_version": result_data.get("pipeline_version", "unknown"),
        "extraction_method": result_data.get("extraction_method", "unknown"),
        "country": result_data.get("detected_sector"),
        "detected_currency": currency_code,
        "currency_confidence": currency_conf,
        "currency_detection_method": currency_method,
        "sector": result_data.get("detected_sector"),
        "duration_months": result_data.get("detected_duration_months"),
        "locations": result_data.get("detected_locations", []),
        "boq_item_count": len(result_data.get("boq_items", []) or []),
        "boq_confidence": result_data.get("boq_confidence"),
        "pricing_available": result_data.get("pricing_result") is not None,
        "workforce_detected": bool(result_data.get("detected_workforce", {})),
        "page_count": result_data.get("metadata", {}).get("page_count", 0),
        "completed_stages": result_data.get("completed_stages", []),
        "failed_stages": result_data.get("failed_stages", []),
        "warning_count": len(result_data.get("warnings", []) or []),
        "engine_versions": {
            "package": PACKAGE_VERSION,
            "pipeline": result_data.get("pipeline_version", "unknown"),
            "extraction": result_data.get("extraction_method", "unknown"),
            "currency_engine": "1.0.0",
            "entity_classifier": "1.0.0",
            "boq_engine": "v2",
            "report_framework": "1.0.0",
            "evidence_system": "1.0.0",
        },
    }

    if evidence_data:
        metadata["evidence_summary"] = {
            "total_fields": evidence_data.get("total_fields", 0),
            "accepted": evidence_data.get("accepted_count", 0),
            "rejected": evidence_data.get("rejected_count", 0),
        }
        audit = evidence_data.get("processing_audit", {})
        metadata["processing_audit"] = {
            "pdf_parsed": audit.get("pdf_parsed", False),
            "page_count": audit.get("page_count", 0),
            "ocr_used": audit.get("ocr_used", False),
            "ocr_reason": audit.get("ocr_reason"),
            "ocr_skipped_reason": audit.get("ocr_skipped_reason"),
            "currency_detected": audit.get("currency_detected"),
            "currency_confidence": audit.get("currency_confidence"),
            "currency_method": audit.get("currency_method"),
            "boq_tables_found": audit.get("boq_tables_found", 0),
            "boq_tables_accepted": audit.get("boq_tables_accepted", 0),
            "boq_tables_rejected": audit.get("boq_tables_rejected", 0),
            "entities_classified": audit.get("numeric_entities_classified", 0),
            "processing_time_ms": audit.get("processing_time_ms"),
            "stages_completed": audit.get("stages_completed", 0),
            "stages_failed": audit.get("stages_failed", 0),
        }

    return metadata


def _format_processing_duration(value: Any) -> str:
    """Format processing duration in milliseconds into a readable string."""
    if value in (None, "", "N/A"):
        return "N/A"
    try:
        total_ms = int(value)
    except (TypeError, ValueError):
        return str(value)

    seconds, milliseconds = divmod(total_ms, 1000)
    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)

    if hours:
        return f"{hours}h {minutes}m {seconds}s"
    if minutes:
        return f"{minutes}m {seconds}s"
    if seconds:
        return f"{seconds}s"
    return f"{milliseconds}ms"


def _format_file_size(size_bytes: int) -> str:
    """Format a byte count for the manifest."""
    units = ["B", "KB", "MB", "GB"]
    size = float(size_bytes)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size_bytes} B"


def _validate_package_entries(
    entries: Dict[str, bytes],
    required_files: List[str],
) -> None:
    """Validate that the ZIP will contain all required, non-empty files."""
    missing = [name for name in required_files if name not in entries]
    if missing:
        raise ValueError(
            "Submission package generation incomplete. Missing required files: "
            + ", ".join(missing)
        )

    empty = [name for name, content in entries.items() if not content]
    if empty:
        raise ValueError(
            "Submission package generation incomplete. Empty files detected: "
            + ", ".join(empty)
        )


def _build_manifest_content(
    job_id: str,
    result_dict: Dict[str, Any],
    evidence_data: Dict[str, Any],
    entries: Dict[str, bytes],
    package_size_bytes: int,
) -> str:
    """Build the package manifest from the actual generated entries."""
    audit = evidence_data.get("processing_audit", {}) or {}
    currency_data = result_dict.get("detected_currency", {})
    currency_code = "Unknown"
    if isinstance(currency_data, dict):
        currency_code = currency_data.get("currency_code") or "Unknown"

    country_detected = audit.get("country_detected") or result_dict.get("jurisdiction") or "Unknown"
    warnings = result_dict.get("warnings", []) or []
    pages_processed = audit.get("page_count") or result_dict.get("metadata", {}).get("page_count") or 0
    tables_extracted = audit.get("boq_tables_accepted", 0)
    evidence_items = evidence_data.get("total_fields", 0)
    processing_status = result_dict.get("status", "unknown")
    package_lines = [
        ("01 Executive Summary.pdf", "Executive Summary.pdf"),
        ("02 Tender Completion Guide.pdf", "Tender Completion Guide.pdf"),
        ("03 Tender Readiness Assessment.pdf", "Tender Readiness Assessment.pdf"),
        ("04 Submission Letter.pdf", "Submission Letter.pdf"),
        ("05 Bid Response Roadmap.pdf", "Bid Response Roadmap.pdf"),
        ("06 Tender Integrity Audit.pdf", "Tender Integrity Audit.pdf"),
        ("07 Processing Audit.pdf", "Processing Audit.pdf"),
        ("08 Evidence Report.pdf", "Evidence Report.pdf"),
        ("09 BOQ.xlsx", "BOQ.xlsx"),
        ("10 BOQ.csv", "BOQ.csv"),
        ("11 Metadata.json", "Metadata.json"),
        ("12 Original Tender.pdf", "Original Tender.pdf (if included)"),
    ]

    contents_lines = []
    for filename, label in package_lines:
        marker = "✓" if filename in entries else "✗"
        contents_lines.append(f"{marker} {label}")

    readiness_report = result_dict.get("readiness_report") if isinstance(result_dict.get("readiness_report"), dict) else build_readiness_report(result_dict)
    readiness_score = ((result_dict.get("readiness_score") or {}).get("overall_score") if isinstance(result_dict.get("readiness_score"), dict) else None) or (((readiness_report or {}).get("readiness_score") or {}).get("overall_score") if isinstance(readiness_report, dict) else None)
    evidence_fields = ((result_dict.get("evidence") or {}).get("fields", {}) or {})
    high_conf = 0
    med_conf = 0
    low_conf = 0
    missing_conf = 0
    for item in evidence_fields.values():
        confidence = str(item.get("confidence") or ("Missing" if item.get("value") in (None, "", [], {}) else "Low")).title()
        if confidence == "High":
            high_conf += 1
        elif confidence == "Medium":
            med_conf += 1
        elif confidence == "Low":
            low_conf += 1
        else:
            missing_conf += 1

    return (
        "Tender Engine\n\n"
        "Submission Package Manifest\n\n"
        "----------------------------------------\n\n"
        f"Package Version\n{PACKAGE_VERSION}\n\n"
        f"Generation Date\n{datetime.now().strftime('%Y-%m-%d')}\n\n"
        f"Generation Time\n{datetime.now().strftime('%H:%M:%S')}\n\n"
        f"Job ID\n{job_id}\n\n"
        f"Pipeline Version\n{result_dict.get('pipeline_version', 'unknown')}\n\n"
        f"Processing Duration\n{_format_processing_duration(audit.get('processing_time_ms'))}\n\n"
        f"Package Size\n{_format_file_size(package_size_bytes)}\n\n"
        "Generated By\nTender Engine\n\n"
        "----------------------------------------\n\n"
        "PACKAGE CONTENTS\n\n"
        + "\n".join(contents_lines)
        + "\n\n----------------------------------------\n\n"
        "PROCESSING SUMMARY\n\n"
        f"Country Detected\n{country_detected}\n\n"
        f"Currency Detected\n{currency_code}\n\n"
        f"Pages Processed\n{pages_processed}\n\n"
        f"OCR Used\n{'Yes' if audit.get('ocr_used') else 'No'}\n\n"
        f"Tables Extracted\n{tables_extracted}\n\n"
        f"Evidence Items\n{evidence_items}\n\n"
        f"Warnings\n{len(warnings)}\n\n"
        f"Processing Status\n{processing_status}\n\n"
        f"Readiness Score\n{readiness_score if readiness_score is not None else 'Unknown'}\n\n"
        "CROSS-REFERENCES\n\n"
        "Completion Guide\nSee 02 Tender Completion Guide.pdf for evidence-based actions and completion steps.\n\n"
        "Readiness Report\nSee 03 Tender Readiness Assessment.pdf for readiness score, deductions, and compliance impacts.\n\n"
        "Submission Letter\nSee 04 Submission Letter.pdf for tender-facing submission details.\n\n"
        "Audit Report\nSee 06 Tender Integrity Audit.pdf for processing and integrity risks.\n\n"
        "Roadmap\nSee 05 Bid Response Roadmap.pdf for submission sequencing.\n\n"
        "EXTRACTION CONFIDENCE SUMMARY\n\n"
        f"High Confidence\n{high_conf}\n\n"
        f"Medium Confidence\n{med_conf}\n\n"
        f"Low Confidence\n{low_conf}\n\n"
        f"Missing\n{missing_conf}\n\n"
        "----------------------------------------\n\n"
        "INTEGRITY\n\n"
        "This package was generated from verified document extraction.\n\n"
        "No information has been invented.\n\n"
        "No missing values have been fabricated.\n\n"
        "Always verify documentation before submission.\n\n"
        "----------------------------------------\n\n"
        "Generated by Tender Engine\n"
    )


def _serialize_zip(entries: Dict[str, bytes], manifest: str) -> bytes:
    """Serialize the package entries plus manifest into a ZIP archive."""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in sorted(entries.keys()):
            zf.writestr(name, entries[name])
        zf.writestr("PACKAGE_MANIFEST.txt", manifest.encode("utf-8"))
        if zf.testzip() is not None:
            raise ValueError("Submission package ZIP failed integrity validation.")

    zip_bytes = zip_buffer.getvalue()
    if not zipfile.is_zipfile(BytesIO(zip_bytes)):
        raise ValueError("Submission package ZIP is invalid.")
    return zip_bytes


# ═══════════════════════════════════════════════════════════════════════
# Build BOQ Excel
# ═══════════════════════════════════════════════════════════════════════

def _build_boq_xlsx(result_data: Dict[str, Any]) -> Optional[BytesIO]:
    """Build BOQ Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("[PACKAGE] openpyxl not available — skipping XLSX")
        return None

    boq_items = result_data.get("boq_items", []) or []
    if not boq_items:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Headers
    headers = ["Item No", "Description", "Quantity", "Unit", "Rate", "Amount", "Section"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data
    for row_idx, item in enumerate(boq_items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("item_no", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=item.get("description", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=item.get("quantity")).border = thin_border
        ws.cell(row=row_idx, column=4, value=item.get("unit", "")).border = thin_border
        ws.cell(row=row_idx, column=5, value=item.get("rate")).border = thin_border
        ws.cell(row=row_idx, column=6, value=item.get("amount")).border = thin_border
        ws.cell(row=row_idx, column=7, value=item.get("section", "")).border = thin_border

    # Auto-width
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════
# Build BOQ CSV
# ═══════════════════════════════════════════════════════════════════════

def _build_boq_csv(result_data: Dict[str, Any]) -> Optional[BytesIO]:
    """Build BOQ CSV file."""
    import csv

    boq_items = result_data.get("boq_items", []) or []
    if not boq_items:
        return None

    buffer = BytesIO()
    text_buffer = io.TextIOWrapper(buffer, encoding='utf-8', newline='')
    writer = csv.writer(text_buffer)

    writer.writerow(["Item No", "Description", "Quantity", "Unit", "Rate", "Amount", "Section"])
    for item in boq_items:
        writer.writerow([
            item.get("item_no", ""),
            item.get("description", ""),
            item.get("quantity"),
            item.get("unit", ""),
            item.get("rate"),
            item.get("amount"),
            item.get("section", ""),
        ])

    text_buffer.flush()
    text_buffer.detach()
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════
# Main Entry Point — Generate Enterprise Submission Package ZIP
# ═══════════════════════════════════════════════════════════════════════

def generate_submission_package_zip(
    job_id: str,
    result_dict: Dict[str, Any],
    company_name_override: Optional[str] = None,
    company_address_override: Optional[str] = None,
    original_pdf_path: Optional[str] = None,
) -> BytesIO:
    """
    Generate a complete enterprise submission package ZIP.

    Package contents:
      01 Executive Summary.pdf
      02 Tender Completion Guide.pdf
      03 Tender Readiness Assessment.pdf
      04 Submission Letter.pdf
      05 Bid Response Roadmap.pdf
      06 Tender Integrity Audit.pdf
      07 Processing Audit.pdf
      08 Evidence Report.pdf
      09 BOQ.xlsx
      10 BOQ.csv
      11 Metadata.json
      12 Original Tender.pdf (if available)
      PACKAGE_MANIFEST.txt

    All reports use common branding, headers, footers, page numbering.
    """
    evidence_data = build_evidence_system(result_dict)
    pdf_styles = get_pdf_styles()
    entries: Dict[str, bytes] = {}

    # ── 01 Executive Summary.pdf ──────────────────────────────────────
    readiness_report = build_readiness_report(result_dict)
    result_dict = dict(result_dict)
    result_dict["readiness_report"] = readiness_report
    report = build_professional_report(result_dict)
    story_01: List[Any] = []
    _build_cover_page(story_01, result_dict, job_id)
    _build_table_of_contents(story_01)
    story_01.extend(build_pdf_sections(report, pdf_styles))
    entries["01 Executive Summary.pdf"] = _build_pdf_document(
        "Executive Summary", story_01, job_id
    ).getvalue()

    # ── 02 / 05 Guide + Roadmap PDFs ─────────────────────────────────
    completion_guide_pdf = generate_completion_guide(job_id, result_dict).getvalue()
    roadmap_pdf = generate_roadmap_pdf(job_id, result_dict).getvalue()
    entries["02 Tender Completion Guide.pdf"] = completion_guide_pdf
    entries["05 Bid Response Roadmap.pdf"] = roadmap_pdf

    # ── 03 Tender Readiness Assessment.pdf ────────────────────────────
    entries["03 Tender Readiness Assessment.pdf"] = generate_readiness_pdf(
        job_id, result_dict
    ).getvalue()

    # ── 04 Submission Letter.pdf ──────────────────────────────────────
    entries["04 Submission Letter.pdf"] = generate_submission_letter(
        job_id,
        result_dict,
        company_name_override=company_name_override,
        company_address_override=company_address_override,
    ).getvalue()

    # ── 06 Tender Integrity Audit.pdf ─────────────────────────────────
    entries["06 Tender Integrity Audit.pdf"] = generate_audit_pdf(
        job_id, result_dict
    ).getvalue()

    # ── 07 / 08 Processing + Evidence PDFs ────────────────────────────
    evidence_story = build_evidence_pdf_sections(evidence_data, pdf_styles)
    entries["07 Processing Audit.pdf"] = _build_pdf_document(
        "Processing Audit", list(evidence_story), job_id
    ).getvalue()
    entries["08 Evidence Report.pdf"] = _build_pdf_document(
        "Evidence Report", list(evidence_story), job_id
    ).getvalue()

    # ── 09 / 10 BOQ exports ───────────────────────────────────────────
    xlsx = _build_boq_xlsx(result_dict)
    if xlsx:
        entries["09 BOQ.xlsx"] = xlsx.getvalue()

    csv_data = _build_boq_csv(result_dict)
    if csv_data:
        entries["10 BOQ.csv"] = csv_data.getvalue()

    # ── 11 Metadata.json ──────────────────────────────────────────────
    metadata = _build_metadata(result_dict, job_id, evidence_data)
    entries["11 Metadata.json"] = json.dumps(
        metadata, indent=2, default=str
    ).encode("utf-8")

    # ── 12 Original Tender.pdf (optional) ─────────────────────────────
    if original_pdf_path and os.path.exists(original_pdf_path):
        with open(original_pdf_path, "rb") as f:
            entries["12 Original Tender.pdf"] = f.read()

    required_files = [
        "01 Executive Summary.pdf",
        "02 Tender Completion Guide.pdf",
        "03 Tender Readiness Assessment.pdf",
        "04 Submission Letter.pdf",
        "05 Bid Response Roadmap.pdf",
        "06 Tender Integrity Audit.pdf",
        "07 Processing Audit.pdf",
        "08 Evidence Report.pdf",
        "11 Metadata.json",
    ]
    _validate_package_entries(entries, required_files)

    if "09 BOQ.xlsx" not in entries or "10 BOQ.csv" not in entries:
        logger.warning(
            "[PACKAGE] BOQ exports unavailable for job %s; package will include manifest markers.",
            job_id,
        )

    manifest = _build_manifest_content(job_id, result_dict, evidence_data, entries, 0)
    zip_bytes = _serialize_zip(entries, manifest)
    final_size = len(zip_bytes)
    manifest = _build_manifest_content(job_id, result_dict, evidence_data, entries, final_size)
    zip_bytes = _serialize_zip(entries, manifest)
    if len(zip_bytes) != final_size:
        manifest = _build_manifest_content(
            job_id, result_dict, evidence_data, entries, len(zip_bytes)
        )
        zip_bytes = _serialize_zip(entries, manifest)

    zip_buffer = BytesIO(zip_bytes)
    zip_buffer.seek(0)
    file_count = len(entries) + 1
    logger.info(
        "[PACKAGE] Enterprise submission package generated for job %s — %d files, %s",
        job_id, file_count, _format_file_size(len(zip_bytes)),
    )
    return zip_buffer


# ═══════════════════════════════════════════════════════════════════════
# Backward Compatibility
# ═══════════════════════════════════════════════════════════════════════

def generate_submission_package(
    job_id: str,
    result_dict: Dict[str, Any],
    company_name_override: Optional[str] = None,
    company_address_override: Optional[str] = None,
) -> BytesIO:
    """
    Generate a comprehensive Submission Package PDF (legacy single-PDF format).
    
    This is kept for backward compatibility. New code should use
    generate_submission_package_zip() for the enterprise ZIP format.
    """
    buffer = BytesIO()
    pdf_styles = get_pdf_styles()
    readiness_report = build_readiness_report(result_dict)
    result_dict = dict(result_dict)
    result_dict["readiness_report"] = readiness_report
    report = build_professional_report(result_dict)

    story: List[Any] = []
    _build_cover_page(story, result_dict, job_id)
    _build_table_of_contents(story)
    story.extend(build_pdf_sections(report, pdf_styles))

    # Add evidence sections
    evidence_data = build_evidence_system(result_dict)
    story.extend(build_evidence_pdf_sections(evidence_data, pdf_styles))

    # Footer
    story.append(Spacer(1, 1 * cm))
    story.append(_hr())
    story.append(Paragraph(
        f"Generated by Tender Engine | Job ID: {job_id[:16]}... | "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')} | v{PACKAGE_VERSION}",
        FOOTER_STYLE,
    ))
    story.append(Paragraph(VERIFICATION_NOTICE_FULL, FOOTER_STYLE))

    pdf = _build_pdf_document("Submission Package", story, job_id)
    buffer = pdf
    logger.info("[PACKAGE] Legacy submission package PDF generated for job %s", job_id)
    return buffer
